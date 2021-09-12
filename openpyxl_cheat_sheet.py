from openpyxl import load_workbook

#official documentation
# https://openpyxl.readthedocs.io/en/stable/tutorial.html

#open file
wb_handle = load_workbook(filename, data_only=False)
sh = wb["Sheet_name"]
sheet_names = self.wb_handle.sheetnames # Get sheet names
self.sheet_handle = self.wb_handle.active

#iteration


#collect cell value
for cell in row:
	cell_value = cell.value


for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
  for cell in row:
    print(cell)


for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
  for cell in col:
    print(cell)

#get specific column
ws['A']